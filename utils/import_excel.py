"""
One-time importer: reads Golf 2026.xlsx and writes structured CSV/JSON
files into the data/ directory.  Safe to re-run (overwrites).

Usage:
    python -m utils.import_excel            (from project root)
    python utils/import_excel.py            (direct)
"""

import json
import sys
from pathlib import Path

import openpyxl
import pandas as pd

# Resolve paths regardless of how the script is invoked
_SCRIPT_DIR = Path(__file__).resolve().parent
_PROJECT_DIR = _SCRIPT_DIR.parent
_DATA_DIR = _PROJECT_DIR / "data"
_EXCEL_PATH = _PROJECT_DIR / "Golf 2026.xlsx"


def _ensure_dirs():
    _DATA_DIR.mkdir(parents=True, exist_ok=True)


# ---------------------------------------------------------------------------
# Goals
# ---------------------------------------------------------------------------

def _import_goals(wb: openpyxl.Workbook) -> dict:
    ws = wb["Goals"]
    goals: dict = {"big_goals": [], "component_goals": [], "sub_goals": {}}

    section = None
    current_sub = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        a_val = row[0].value
        b_val = row[1].value if len(row) > 1 else None

        # Detect section headers
        if a_val == "Big Goals":
            section = "big"
            continue
        elif a_val == "Component Goals":
            section = "component"
            continue
        elif a_val == "Sub Goals":
            section = "sub"
            continue

        # Sub-goal group headers like "Goal 1 (Driver)"
        if section == "sub" and a_val and isinstance(a_val, str) and a_val.startswith("Goal"):
            current_sub = a_val
            goals["sub_goals"][current_sub] = []
            continue

        # Data rows
        if section == "big" and b_val:
            goals["big_goals"].append(str(b_val))
        elif section == "component" and b_val:
            goals["component_goals"].append(str(b_val))
        elif section == "sub" and current_sub and b_val:
            goals["sub_goals"][current_sub].append(str(b_val))

    return goals


# ---------------------------------------------------------------------------
# Practice logs  (ball_striking, putting, short_game)
# ---------------------------------------------------------------------------

_BALL_STRIKING_COLS = [
    "date",
    "mechanical_no_results",
    "towel_drill_3x",
    "eyes_close_strike_3x",
    "toe_heel_center_3x",
    "jump_the_ball",
    "wedge_ladder_3x",
    "crazy_shit_1x",
    "one_handed_pitch_3x",
]

_PUTTING_COLS = [
    "date",
    "three_foot_drill",
    "guess_the_slope",
    "lag_drill",
]

_SHORT_GAME_COLS = [
    "date",
]


def _import_practice_sheet(wb, sheet_name, columns) -> pd.DataFrame:
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue
        record: dict = {}
        for i, col_name in enumerate(columns):
            val = row[i] if i < len(row) else None
            if col_name == "date" and val is not None:
                val = pd.Timestamp(val).strftime("%Y-%m-%d")
            record[col_name] = val
        rows.append(record)
    if not rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(rows, columns=columns)


# ---------------------------------------------------------------------------
# Drill descriptions
# ---------------------------------------------------------------------------

def _import_drills(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb["Description"]
    drills = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
        name = row[0].value
        if not name:
            continue
        levels = {}
        for i, level_label in enumerate(["Level 1", "Level 2", "Level 3", "Level 4"], start=1):
            val = row[i].value if i < len(row) else None
            if val:
                levels[level_label] = str(val)
        desc = row[6].value if len(row) > 6 else None
        drills.append({
            "name": str(name),
            "levels": levels,
            "description": str(desc) if desc else None,
        })
    return drills


# ---------------------------------------------------------------------------
# Testing lookup tables  &  test results
# ---------------------------------------------------------------------------

def _import_testing(wb: openpyxl.Workbook) -> tuple[dict, pd.DataFrame]:
    ws = wb["Testing"]

    # --- Lookup tables (columns P onward, rows 2-9) ---
    lookup: dict[str, list[dict]] = {}
    # Row 1 has the index numbers 1..27 in Q1..AQ1
    indices = []
    for col_idx in range(17, ws.max_column + 1):  # Q=17 (1-indexed)
        val = ws.cell(row=1, column=col_idx).value
        if val is not None:
            indices.append(int(val))

    for row_idx in range(2, 10):  # rows 2-9
        shot_name = ws.cell(row=row_idx, column=16).value  # column P
        if not shot_name:
            continue
        handicaps = []
        for i, col_idx in enumerate(range(17, 17 + len(indices))):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                handicaps.append({"score": indices[i], "handicap": float(val)})
        lookup[str(shot_name)] = handicaps

    # --- Test results (columns A-M, row 3 onward) ---
    test_cols = [
        "date", "total", "avg_handicap", "adj", "handicap",
        "50_yards_f", "30_yards_f", "10_f_chip", "20_yards_r",
        "flop", "15_f_pitch", "8_yard_sand", "15_yard_sand",
    ]
    results = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue
        record = {}
        for i, col_name in enumerate(test_cols):
            val = row[i] if i < len(row) else None
            if col_name == "date" and val is not None:
                val = pd.Timestamp(val).strftime("%Y-%m-%d")
            elif isinstance(val, str) and val.startswith("="):
                val = None  # skip formulas
            elif val == "na":
                val = None
            record[col_name] = val
        results.append(record)

    results_df = pd.DataFrame(results, columns=test_cols) if results else pd.DataFrame(columns=test_cols)
    return lookup, results_df


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run_import(excel_path=None):
    excel_path = excel_path or _EXCEL_PATH
    if not excel_path.exists():
        print(f"Excel file not found: {excel_path}")
        sys.exit(1)

    _ensure_dirs()
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)

    # Goals
    goals = _import_goals(wb)
    with open(_DATA_DIR / "goals.json", "w") as f:
        json.dump(goals, f, indent=2)
    print(f"  goals.json  ({len(goals['big_goals'])} big, "
          f"{len(goals['component_goals'])} component, "
          f"{sum(len(v) for v in goals['sub_goals'].values())} sub-goals)")

    # Practice logs
    bs_df = _import_practice_sheet(wb, "ball_striking_practice", _BALL_STRIKING_COLS)
    bs_df.to_csv(_DATA_DIR / "ball_striking.csv", index=False)
    print(f"  ball_striking.csv  ({len(bs_df)} sessions)")

    put_df = _import_practice_sheet(wb, "putting_practice", _PUTTING_COLS)
    put_df.to_csv(_DATA_DIR / "putting.csv", index=False)
    print(f"  putting.csv  ({len(put_df)} sessions)")

    sg_df = _import_practice_sheet(wb, "short_game_practice", _SHORT_GAME_COLS)
    sg_df.to_csv(_DATA_DIR / "short_game.csv", index=False)
    print(f"  short_game.csv  ({len(sg_df)} sessions)")

    # Drills
    drills = _import_drills(wb)
    with open(_DATA_DIR / "drills.json", "w") as f:
        json.dump(drills, f, indent=2)
    print(f"  drills.json  ({len(drills)} drills)")

    # Testing
    lookup, test_df = _import_testing(wb)
    with open(_DATA_DIR / "testing_lookup.json", "w") as f:
        json.dump(lookup, f, indent=2)
    print(f"  testing_lookup.json  ({len(lookup)} shot types)")

    test_df.to_csv(_DATA_DIR / "testing.csv", index=False)
    print(f"  testing.csv  ({len(test_df)} test sessions)")

    wb.close()
    print("\nImport complete!")


if __name__ == "__main__":
    print("Importing Golf 2026.xlsx ...")
    run_import()
